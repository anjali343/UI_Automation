import inspect
import logging

import pytest
import openpyxl


@pytest.mark.usefixtures("setup")
class BaseClass:

    def getLogger(self):
        loggerName = inspect.stack()[1][3]
        logger = logging.getLogger(loggerName)
        fileHandler = logging.FileHandler('logfile.log')
        formatter = logging.Formatter("%(asctime)s :%(levelname)s : %(name)s :%(message)s")
        fileHandler.setFormatter(formatter)

        logger.addHandler(fileHandler)  # filehandler object

        logger.setLevel(logging.DEBUG)
        return logger

    def getRowCount(self, file, sheetName):
        """:return the maximun row count from Excel data sheet"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return(sheet.max_row)

    def getColumnCount(self, file, sheetName):
        """return the maximum column count from Excel data sheet."""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return (sheet.max_column)

    def readData(self, file, sheetName, rownum, columnno):
        """read data from particular row and column and return value"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.cell(row = rownum, column = columnno).value

    def writeData(self, file, sheetName, rownum, columnno, data):
        """ writes the data back to Excel file and save it"""
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        sheet.cell(row=rownum, column = columnno).value = data
        workbook.save(file)

    def formDictionary(self, file, sheetName):
        """return the list the form for nested list and dictionary ."""
        data = []
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        rows = self.getRowCount(file, sheetName)
        columns = self.getColumnCount(file, sheetName)
        for i in range(2, rows + 1): # to get rows
            Dict = {}
            for j in range(1, columns + 1):  # to get columns
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data.append(Dict)
        return data



